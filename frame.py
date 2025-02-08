from tkinter import Tk, Frame, Label, Button
from tkinter.filedialog import asksaveasfilename
from tkcalendar import DateEntry
from logger import logger
import load_configuration
import connectionSQL as cs
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dataclasses import dataclass
from typing import List, Dict, Tuple

@dataclass
class ExcelStyles:
    """Clase para manejar los estilos de Excel."""
    bold_font: Font = Font(bold=True, size=12, color="000000")
    header_fill: PatternFill = PatternFill(start_color="FFA07A", end_color="FFA07A", fill_type="solid")
    fill_light: PatternFill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_dark: PatternFill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    center_alignment: Alignment = Alignment(horizontal="center", vertical="center")
    border: Border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

class DatabaseManager:
    """Clase para manejar las operaciones de base de datos."""
    
    def __init__(self):
        self.configuration = load_configuration.configuration()
        self.param = self.configuration.cargar_configuracion()
        
    def execute_queries(self, fecha_desde: str, fecha_hasta: str) -> Tuple[List, List]:
        """Ejecuta las consultas para obtener datos evadidos y no evadidos."""
        try:
            base_query = (
                "SELECT CONCAT(RTRIM(seriefolio), numcheque), fecha, cierre, total, fp.descripcion "
                "FROM cheques c "
                "INNER JOIN chequespagos cp ON c.folio = cp.folio "
                "INNER JOIN formasdepago fp ON cp.idformadepago = fp.idformadepago "
                f"WHERE fecha BETWEEN '{fecha_desde}' AND '{fecha_hasta}' AND "
            )
            
            connection = cs.SQLServerConnection(
                self.param.server,
                self.param.database,
                self.param.username,
                self.param.password,
                self.param.port
            )
            connection.connect()
            query_executor = cs.SQLServerQueryExecutor(connection.connection)
            
            evadidos = query_executor.execute_query(base_query + "campoadicional3 = '*'")
            no_evadidos = query_executor.execute_query(base_query + "campoadicional3 != '*'")
            
            connection.disconnect()
            return evadidos, no_evadidos
            
        except Exception as e:
            logger.write_to_log(f"Error en DatabaseManager.execute_queries: {str(e)}")
            raise

class ExcelReportGenerator:
    """Clase para generar reportes en Excel."""
    
    def __init__(self):
        self.styles = ExcelStyles()
        self.column_widths = {
            "B": 21.57,  # Apertura
            "C": 21.57,  # Cierre
            "D": 21.57,  # Total
            "E": 21.57,  # Forma de pago
        }
    
    def format_worksheet(self, ws, datos: List) -> None:
        """Aplica formato a una hoja de trabajo."""
        try:
            encabezados = ["Folio", "Apertura", "Cierre", "Total", "Forma de pago"]
            ws.append(encabezados)
            
            # Formato de encabezados
            for cell in ws[1]:
                self._apply_header_style(cell)
            
            # Procesar datos y calcular totales
            totales_pago = {}
            total_suma = self._process_data_rows(ws, datos, totales_pago)
            
            # Agregar totales
            self._add_total_row(ws, len(datos) + 2, total_suma)
            self._add_payment_summary(ws, totales_pago)
            
            # Ajustar anchos de columna
            self._adjust_column_widths(ws)
            
        except Exception as e:
            logger.write_to_log(f"Error en format_worksheet: {str(e)}")
            raise
    
    def _apply_header_style(self, cell) -> None:
        """Aplica estilo a una celda de encabezado."""
        cell.font = self.styles.bold_font
        cell.fill = self.styles.header_fill
        cell.alignment = self.styles.center_alignment
        cell.border = self.styles.border
    
    def _process_data_rows(self, ws, datos: List, totales_pago: Dict) -> float:
        """Procesa las filas de datos y calcula totales."""
        total_suma = 0
        for index, row in enumerate(datos, start=2):
            ws.append(list(row))
            fill_color = self.styles.fill_light if index % 2 == 0 else self.styles.fill_dark
            
            for col_idx, cell in enumerate(ws[index], 1):
                self._format_data_cell(cell, col_idx, fill_color)
                
                if col_idx == 4:  # Columna Total
                    total_suma += float(cell.value)
                    self._update_payment_totals(totales_pago, ws[index][4].value, float(cell.value))
        
        return total_suma
    
    def _format_data_cell(self, cell, col_idx: int, fill_color: PatternFill) -> None:
        """Aplica formato a una celda de datos."""
        cell.alignment = self.styles.center_alignment
        cell.border = self.styles.border
        cell.fill = fill_color
        
        if col_idx == 4:  # Columna Total
            cell.number_format = '"$" #,##0.00'
    
    def _update_payment_totals(self, totales_pago: Dict, forma_pago: str, valor: float) -> None:
        """Actualiza los totales por forma de pago."""
        if forma_pago not in totales_pago:
            totales_pago[forma_pago] = 0
        totales_pago[forma_pago] += valor
    
    def _add_total_row(self, ws, row: int, total: float) -> None:
        """Agrega la fila de total general."""
        ws[f"C{row}"] = "Total General:"
        ws[f"D{row}"] = total
        
        for col in ["C", "D"]:
            cell = ws[f"{col}{row}"]
            cell.font = self.styles.bold_font
            cell.alignment = self.styles.center_alignment
            cell.fill = self.styles.header_fill
        
        ws[f"D{row}"].number_format = '"$" #,##0.00'
    
    def _add_payment_summary(self, ws, totales_pago: Dict) -> None:
        """Agrega el resumen de pagos."""
        ws["G1"] = "Forma de pago"
        ws["H1"] = "Total"
        
        for col in ["G", "H"]:
            self._apply_header_style(ws[f"{col}1"])
        
        for i, (forma_pago, total) in enumerate(totales_pago.items(), start=2):
            fill = self.styles.fill_light if i % 2 == 0 else self.styles.fill_dark
            
            ws[f"G{i}"] = forma_pago
            ws[f"H{i}"] = total
            
            for col in ["G", "H"]:
                cell = ws[f"{col}{i}"]
                cell.alignment = self.styles.center_alignment
                cell.border = self.styles.border
                cell.fill = fill
            
            ws[f"H{i}"].number_format = '"$" #,##0.00'
    
    def _adjust_column_widths(self, ws) -> None:
        """Ajusta el ancho de las columnas."""
        for col, width in self.column_widths.items():
            ws.column_dimensions[col].width = width
        ws.column_dimensions["G"].width = 21
        ws.column_dimensions["H"].width = 22

class ReporteGUI:
    """Clase para manejar la interfaz gráfica."""
    
    def __init__(self):
        self.root = None
        self.desde_date = None
        self.hasta_date = None
        self.db_manager = DatabaseManager()
        self.excel_generator = ExcelReportGenerator()
    
    def setup_window(self) -> None:
        """Configura la ventana principal."""
        self.root = Tk()
        self.root.withdraw()
        self.root.title("Reporte de cuentas")
        
        try:
            self.root.iconbitmap("logo_sr.ico")
        except Exception as e:
            logger.write_to_log(f"Error setting icon: {str(e)}")
        
        self._center_window(280, 180)
        self.root.resizable(False, False)
    
    def _center_window(self, width: int, height: int) -> None:
        """Centra la ventana en la pantalla."""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        self.root.geometry(f"{width}x{height}+{x}+{y}")
    
    def create_widgets(self) -> None:
        """Crea los widgets de la interfaz."""
        frame = Frame(self.root, padx=20, pady=20)
        frame.pack(expand=True, fill="both")
        
        # Fecha desde
        Label(frame, text="Desde:", anchor="w").grid(row=0, column=0, sticky="w", pady=2)
        self.desde_date = DateEntry(frame, date_pattern='dd/mm/yyyy', locale='es_ES')
        self.desde_date.grid(row=1, column=0, sticky="ew", pady=2)
        
        # Fecha hasta
        Label(frame, text="Hasta:", anchor="w").grid(row=2, column=0, sticky="w", pady=2)
        self.hasta_date = DateEntry(frame, date_pattern='dd/mm/yyyy', locale='es_ES')
        self.hasta_date.grid(row=3, column=0, sticky="ew", pady=2)
        
        # Botón aceptar
        Button(frame, text="Aceptar", command=self._generar_reporte).grid(
            row=4, column=0, sticky="e", pady=15
        )
        
        frame.columnconfigure(0, weight=1)
    
    def _generar_reporte(self) -> None:
        """Genera el reporte Excel."""
        try:
            evadidos, no_evadidos = self.db_manager.execute_queries(
                self.desde_date.get(),
                self.hasta_date.get()
            )
            
            wb = openpyxl.Workbook()
            
            # Hoja de evadidos
            ws_evadidos = wb.active
            ws_evadidos.title = "Evadidos"
            self.excel_generator.format_worksheet(ws_evadidos, evadidos)
            
            # Hoja de no evadidos
            ws_no_evadidos = wb.create_sheet(title="No evadidos")
            self.excel_generator.format_worksheet(ws_no_evadidos, no_evadidos)
            
            # Guardar archivo
            filename = asksaveasfilename(
                title="Guardar reporte",
                defaultextension=".xlsx",
                initialfile="ReporteCuentas.xlsx",
                filetypes=[("Archivos de Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
            )
            
            if filename:
                wb.save(filename)
                
        except Exception as e:
            logger.write_to_log(f"Error en generar_reporte: {str(e)}")
    
    def run(self) -> None:
        """Inicia la aplicación."""
        try:
            self.setup_window()
            self.create_widgets()
            self.root.deiconify()
            self.root.mainloop()
        except Exception as e:
            logger.write_to_log(f"Error running application: {str(e)}")

def main():
    """Función principal."""
    try:
        app = ReporteGUI()
        app.run()
    except Exception as e:
        logger.write_to_log(f"Error in main: {str(e)}")

if __name__ == "__main__":
    main()